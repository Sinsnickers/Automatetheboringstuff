import openpyxl,os

os.chdir("Chapter13")
wb = openpyxl.Workbook()
print(wb.sheetnames)
print("-------------------------------")
wb.create_sheet()
print(wb.sheetnames)
print("-------------------------------")
wb.create_sheet(index = 0, title = "First sheet")
print(wb.sheetnames)
print("-------------------------------")
wb.create_sheet(index = 2, title = "Middle sheet")
print(wb.sheetnames)
print("-------------------------------")
del wb["Middle sheet"]
del wb["Sheet1"]
print(wb.sheetnames)
sheet = wb["Sheet"]
sheet["A1"] = "Hello Al"
print(sheet["A1"].value)
#wb.save("example_copy2.xlsx")
