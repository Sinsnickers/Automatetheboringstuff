#! python3
# updateProduce.py - Corrects price in produce sales spreadsheet

import openpyxl, os

os.chdir("Chapter13\\ChapterProjects")

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]

#The produce types and their updates

PRICE_UPDATES = {"Garlic": 3.07,
                "Celery" : 1.19,
                "Lemon" : 1.27}

# TODO: Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row = rowNum, column = 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

#wb.save("produceSales_copy.xlsx")