import openpyxl, os, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

try:
    userInput = int(sys.argv[1])
except:
    print("Please enter a valid digit")
#change current working directory
#os.chdir("Chapter13\\PracticeProjects")

#open new Excel file
wb = openpyxl.Workbook()
sheet = wb.active

#freeze the first row and column
sheet.freeze_panes = "B2"
for i in range(1,int(userInput)+1):
    sheet["A"+str(i+1)].font = Font(bold = True)
    sheet["A"+str(i+1)] = i
    sheet[get_column_letter(i+1)+str(1)] = i
    sheet[get_column_letter(i+1)+str(1)].font = Font(bold = True)
    for column in range(2,int(userInput)+2):
        sheet[get_column_letter(i+1)+str(column)] = i*int(column-1)

wb.save(f"multiplicationTable{userInput}.xlsx")

