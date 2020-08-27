import openpyxl, os 
from openpyxl.utils import get_column_letter, column_index_from_string

os.chdir("automate_online-materials")

wb = openpyxl.load_workbook("example.xlsx")

sheet = wb["Sheet1"]
cell = sheet.cell(row=1, column=2)
for i in range(1,8,2):
    print(f"{i} --> {sheet.cell(row=i,column=2).value}")
print("-------------------------------------------")
rowMax = sheet.max_row
columnMax = sheet.max_column
print(rowMax)
print(columnMax)
print("-------------------------------------------")
print(get_column_letter(1)) 
print(get_column_letter(27)) 
print(get_column_letter(900)) 
print(get_column_letter(rowMax))
print("-------------------------------------------")
print(column_index_from_string('A')) # Get A's number.
print(column_index_from_string('AA'))




